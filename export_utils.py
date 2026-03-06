from __future__ import annotations

import sqlite3
from io import BytesIO
from pathlib import Path
from typing import Mapping

from audit_utils import ensure_schema
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ALL_PROJECTS = "All Projects"
HEADER_FILL = PatternFill("solid", fgColor="D9E2EF")
HEADER_FONT = Font(bold=True, color="23384D")


def export_excel_report(db_path: Path, month: str | None = None, project: str | None = None) -> bytes:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        _ensure_table(conn)
        all_entries = _fetch_all_entries(conn)
        filtered_entries = _fetch_filtered_entries(conn, month, project)
        monthly_summary = _fetch_monthly_summary(conn)
        return export_excel_report_from_data(
            all_entries=[dict(row) for row in all_entries],
            filtered_entries=[dict(row) for row in filtered_entries],
            monthly_summary=monthly_summary,
            month=month,
            project=project,
        )
    finally:
        conn.close()


def export_excel_report_from_data(
    all_entries: list[Mapping[str, object]],
    filtered_entries: list[Mapping[str, object]],
    monthly_summary: list[list],
    month: str | None = None,
    project: str | None = None,
) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)

    project_summary = _summarize_filtered(filtered_entries, "project_name")
    employee_summary = _summarize_filtered(filtered_entries, "employee_name")

    _add_sheet(
        workbook,
        "All Entries",
        ["Date", "Employee", "Project", "Task", "Hours"],
        [[row["work_date"], row["employee_name"], row["project_name"], row["task_name"], row["hours_spent"]] for row in all_entries],
    )
    _add_sheet(
        workbook,
        "Current View",
        ["Date", "Employee", "Project", "Task", "Hours"],
        [[row["work_date"], row["employee_name"], row["project_name"], row["task_name"], row["hours_spent"]] for row in filtered_entries],
    )
    _add_statistics_sheet(workbook, month, project, project_summary, employee_summary)
    _add_sheet(workbook, "Monthly Summary", ["Month", "Entries", "Total Hours"], monthly_summary)

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _ensure_table(conn: sqlite3.Connection) -> None:
    ensure_schema(conn)


def _fetch_all_entries(conn: sqlite3.Connection) -> list[sqlite3.Row]:
    return conn.execute(
        "SELECT work_date, employee_name, project_name, task_name, hours_spent FROM work_entries ORDER BY work_date DESC, id DESC"
    ).fetchall()


def _fetch_filtered_entries(conn: sqlite3.Connection, month: str | None, project: str | None) -> list[sqlite3.Row]:
    query = "SELECT work_date, employee_name, project_name, task_name, hours_spent FROM work_entries WHERE 1 = 1"
    params: list[str] = []
    if month:
        query += " AND substr(work_date, 1, 7) = ?"
        params.append(month)
    if project and project != ALL_PROJECTS:
        query += " AND project_name = ?"
        params.append(project)
    query += " ORDER BY work_date DESC, id DESC"
    return conn.execute(query, params).fetchall()


def _fetch_monthly_summary(conn: sqlite3.Connection) -> list[list]:
    rows = conn.execute(
        """
        SELECT substr(work_date, 1, 7) AS month_key, COUNT(*) AS entry_count, ROUND(SUM(hours_spent), 2) AS total_hours
        FROM work_entries
        GROUP BY substr(work_date, 1, 7)
        ORDER BY month_key DESC
        """
    ).fetchall()
    return [[row["month_key"], row["entry_count"], row["total_hours"]] for row in rows]


def _summarize_filtered(rows: list[Mapping[str, object]], key: str) -> list[list]:
    summary: dict[str, list[float]] = {}
    for row in rows:
        label = str(row[key])
        if label not in summary:
            summary[label] = [0, 0.0]
        summary[label][0] += 1
        summary[label][1] += float(row["hours_spent"])
    ordered = sorted(summary.items(), key=lambda item: (-item[1][1], item[0].lower()))
    return [[label, values[0], round(values[1], 2)] for label, values in ordered]


def _add_sheet(workbook: Workbook, title: str, headers: list[str], rows: list[list]) -> None:
    worksheet = workbook.create_sheet(title=title)
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    _fit_columns(worksheet)


def _add_statistics_sheet(
    workbook: Workbook,
    month: str | None,
    project: str | None,
    project_summary: list[list],
    employee_summary: list[list],
) -> None:
    worksheet = workbook.create_sheet(title="Statistics")
    worksheet["A1"] = "Working Hours Statistics"
    worksheet["A1"].font = Font(bold=True, size=14, color="1F3A5F")
    worksheet["A3"] = "Month Filter"
    worksheet["B3"] = month or "All Months"
    worksheet["A4"] = "Project Filter"
    worksheet["B4"] = project or ALL_PROJECTS
    worksheet["A6"] = "Project Summary"
    worksheet["A10"] = "Employee Summary"
    worksheet["A6"].font = Font(bold=True, color="23384D")
    worksheet["A10"].font = Font(bold=True, color="23384D")

    project_start = 7
    employee_start = 11
    _write_table(worksheet, project_start, ["Project", "Entries", "Total Hours"], project_summary)
    _write_table(worksheet, employee_start, ["Employee", "Entries", "Total Hours"], employee_summary)

    if project_summary:
        bar_chart = BarChart()
        bar_chart.type = "col"
        bar_chart.style = 10
        bar_chart.title = "Hours by Project"
        bar_chart.y_axis.title = "Hours"
        bar_chart.x_axis.title = "Project"
        data = Reference(worksheet, min_col=3, min_row=project_start, max_row=project_start + len(project_summary))
        categories = Reference(worksheet, min_col=1, min_row=project_start + 1, max_row=project_start + len(project_summary))
        bar_chart.add_data(data, titles_from_data=True)
        bar_chart.set_categories(categories)
        bar_chart.height = 8
        bar_chart.width = 14
        worksheet.add_chart(bar_chart, "E6")

    if employee_summary:
        pie_chart = PieChart()
        pie_chart.title = "Hours Share by Employee"
        data = Reference(worksheet, min_col=3, min_row=employee_start, max_row=employee_start + len(employee_summary))
        labels = Reference(worksheet, min_col=1, min_row=employee_start + 1, max_row=employee_start + len(employee_summary))
        pie_chart.add_data(data, titles_from_data=True)
        pie_chart.set_categories(labels)
        pie_chart.height = 8
        pie_chart.width = 14
        worksheet.add_chart(pie_chart, "E23")

    _fit_columns(worksheet)


def _write_table(worksheet, start_row: int, headers: list[str], rows: list[list]) -> None:
    for col_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=start_row, column=col_index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    row_pointer = start_row + 1
    if rows:
        for row in rows:
            for col_index, value in enumerate(row, start=1):
                worksheet.cell(row=row_pointer, column=col_index, value=value)
            row_pointer += 1
    else:
        worksheet.cell(row=row_pointer, column=1, value="No data")


def _fit_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        lengths = [len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells]
        max_length = min(max(lengths, default=0) + 2, 48)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = max(12, max_length)
