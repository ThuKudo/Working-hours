import json
import os
from collections import defaultdict
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

from flask import Flask, Response, jsonify, request
from openpyxl import load_workbook

from audit_utils import default_operator
from data_backend import (
    create_entry,
    delete_entry,
    fetch_all_entries,
    fetch_entries,
    fetch_filters,
    fetch_history_rows,
    fetch_monthly_summary,
    import_entries,
    init_backend,
    update_entry,
)
from export_utils import export_excel_report_from_data


ALL_PROJECTS = "All Projects"
BASE_DIR = Path(__file__).resolve().parent
DB_PATH = (Path("/tmp") / "worklog.db") if os.environ.get("VERCEL") else (BASE_DIR / "worklog.db")


app = Flask(__name__)
init_backend(DB_PATH)


def build_summary(entries: list[dict]) -> dict:
    by_project: dict[str, float] = defaultdict(float)
    by_employee: dict[str, float] = defaultdict(float)
    by_date: dict[str, float] = defaultdict(float)
    total_hours = 0.0
    for entry in entries:
        hours = float(entry["hours_spent"])
        total_hours += hours
        by_project[entry["project_name"]] += hours
        by_employee[entry["employee_name"]] += hours
        by_date[entry["work_date"]] += hours
    top_project = ""
    if by_project:
        name, hours = max(by_project.items(), key=lambda item: item[1])
        top_project = f"{name} ({hours:.1f}h)"
    return {
        "total_hours": total_hours,
        "entry_count": len(entries),
        "top_project": top_project,
        "by_project": dict(sorted(by_project.items(), key=lambda item: item[1], reverse=True)),
        "by_employee": dict(sorted(by_employee.items(), key=lambda item: item[0].lower())),
        "by_date": dict(sorted(by_date.items())),
    }


def validate_payload(payload: dict) -> tuple[bool, str]:
    required = ["employee_name", "work_date", "project_name", "task_name", "hours_spent", "operator_name"]
    if not all(str(payload.get(field, "")).strip() for field in required):
        return False, "Please complete all fields."
    try:
        datetime.strptime(str(payload["work_date"]), "%Y-%m-%d")
    except ValueError:
        return False, "Work Date must use the YYYY-MM-DD format."
    try:
        hours = float(payload["hours_spent"])
        if hours <= 0:
            raise ValueError
    except (TypeError, ValueError):
        return False, "Hours must be a positive number."
    return True, ""


def _normalize_header(value: object) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def _parse_import_rows(file_bytes: bytes) -> tuple[list[dict], str]:
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True, read_only=True)
    sheet = workbook["All Entries"] if "All Entries" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration:
        return [], "Excel file is empty."

    header_map = {_normalize_header(name): idx for idx, name in enumerate(headers)}
    aliases = {
        "work_date": ["work_date", "date"],
        "employee_name": ["employee_name", "employee"],
        "project_name": ["project_name", "project"],
        "task_name": ["task_name", "task"],
        "hours_spent": ["hours_spent", "hours"],
    }
    col_idx: dict[str, int] = {}
    for key, keys in aliases.items():
        matched = next((header_map[k] for k in keys if k in header_map), None)
        if matched is None:
            return [], f"Missing required column: {key}"
        col_idx[key] = matched

    parsed: list[dict] = []
    errors: list[str] = []
    for row_number, row in enumerate(rows, start=2):
        raw = {k: row[v] if v < len(row) else None for k, v in col_idx.items()}
        if not any(value not in (None, "") for value in raw.values()):
            continue
        work_date_value = raw["work_date"]
        if isinstance(work_date_value, datetime):
            work_date = work_date_value.strftime("%Y-%m-%d")
        elif isinstance(work_date_value, date):
            work_date = work_date_value.isoformat()
        else:
            work_date = str(work_date_value or "").strip()
        employee_name = str(raw["employee_name"] or "").strip()
        project_name = str(raw["project_name"] or "").strip()
        task_name = str(raw["task_name"] or "").strip()
        hours_raw = raw["hours_spent"]

        try:
            datetime.strptime(work_date, "%Y-%m-%d")
            hours = round(float(hours_raw), 2)
            if hours <= 0 or not employee_name or not project_name or not task_name:
                raise ValueError
        except (TypeError, ValueError):
            errors.append(f"Invalid data at row {row_number}.")
            continue

        parsed.append(
            {
                "employee_name": employee_name,
                "work_date": work_date,
                "project_name": project_name,
                "task_name": task_name,
                "hours_spent": hours,
            }
        )

    if errors:
        return [], " ".join(errors[:5])
    if not parsed:
        return [], "No valid rows found in Excel file."
    return parsed, ""


@app.get("/")
def index() -> Response:
    html = HTML_PAGE.replace("__DEFAULT_OPERATOR__", json.dumps(default_operator())[1:-1])
    return Response(html, mimetype="text/html; charset=utf-8")


@app.get("/api/filters")
def api_filters():
    return jsonify(fetch_filters(DB_PATH))


@app.get("/api/dashboard")
def api_dashboard():
    month = request.args.get("month", "").strip()
    if not month:
        month = date.today().strftime("%Y-%m")
    project = request.args.get("project") or ALL_PROJECTS
    entries = fetch_entries(DB_PATH, month, project)
    history = fetch_history_rows(DB_PATH, month=month, project=project, limit=40)
    return jsonify({"entries": entries, "summary": build_summary(entries), "history": history})


@app.post("/api/entries")
def api_create_entry():
    payload = request.get_json(silent=True) or {}
    ok, error = validate_payload(payload)
    if not ok:
        return jsonify({"error": error}), 400

    operator = str(payload["operator_name"]).strip()
    values = {
        "employee_name": str(payload["employee_name"]).strip(),
        "work_date": str(payload["work_date"]).strip(),
        "project_name": str(payload["project_name"]).strip(),
        "task_name": str(payload["task_name"]).strip(),
        "hours_spent": round(float(payload["hours_spent"]), 2),
    }

    create_entry(DB_PATH, values, operator)
    return jsonify({"status": "ok"}), 201


@app.put("/api/entries/<int:entry_id>")
def api_update_entry(entry_id: int):
    payload = request.get_json(silent=True) or {}
    ok, error = validate_payload(payload)
    if not ok:
        return jsonify({"error": error}), 400

    operator = str(payload["operator_name"]).strip()
    values = {
        "employee_name": str(payload["employee_name"]).strip(),
        "work_date": str(payload["work_date"]).strip(),
        "project_name": str(payload["project_name"]).strip(),
        "task_name": str(payload["task_name"]).strip(),
        "hours_spent": round(float(payload["hours_spent"]), 2),
    }

    if not update_entry(DB_PATH, entry_id, values, operator):
        return jsonify({"error": "Entry not found."}), 404
    return jsonify({"status": "ok"})


@app.delete("/api/entries/<int:entry_id>")
def api_delete_entry(entry_id: int):
    payload = request.get_json(silent=True) or {}
    operator = str(payload.get("operator_name", "")).strip()
    if not operator:
        return jsonify({"error": "Please enter the operator name."}), 400

    if not delete_entry(DB_PATH, entry_id, operator):
        return jsonify({"error": "Entry not found."}), 404
    return jsonify({"status": "ok"})


@app.post("/api/import.xlsx")
def api_import_excel():
    operator = str(request.form.get("operator_name", "")).strip()
    if not operator:
        return jsonify({"error": "Please enter the operator name."}), 400
    upload = request.files.get("file")
    if upload is None or not upload.filename:
        return jsonify({"error": "Please choose an Excel file (.xlsx)."}), 400
    if not upload.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "Only .xlsx files are supported."}), 400
    try:
        entries, error = _parse_import_rows(upload.read())
    except Exception:
        return jsonify({"error": "Could not read Excel file."}), 400
    if error:
        return jsonify({"error": error}), 400

    overwrite = str(request.form.get("overwrite", "1")).strip().lower() in {"1", "true", "yes", "on"}
    result = import_entries(DB_PATH, entries, operator, overwrite=overwrite)
    return jsonify({"status": "ok", "result": result})


@app.get("/api/export.xlsx")
def api_export_excel():
    month = request.args.get("month", "").strip()
    if not month:
        month = date.today().strftime("%Y-%m")
    project = request.args.get("project") or ALL_PROJECTS
    all_entries = fetch_all_entries(DB_PATH)
    filtered_entries = fetch_entries(DB_PATH, month, project)
    monthly_summary = fetch_monthly_summary(DB_PATH)
    file_bytes = export_excel_report_from_data(
        all_entries=all_entries,
        filtered_entries=filtered_entries,
        monthly_summary=monthly_summary,
        month=month,
        project=project,
    )
    file_name = f"working-hours-report-{date.today().isoformat()}.xlsx"
    return Response(
        file_bytes,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{file_name}"'},
    )


# =============================================================================
# MODERN DASHBOARD HTML — Full Redesign
# =============================================================================
HTML_PAGE = r"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8"/>
  <meta name="viewport" content="width=device-width,initial-scale=1.0"/>
  <title>ThuKudo — Time Tracker</title>
  <meta name="description" content="Professional working hours tracker and timesheet dashboard by ThuKudo"/>
  <link rel="preconnect" href="https://fonts.googleapis.com"/>
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap" rel="stylesheet"/>
  <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.7/dist/chart.umd.min.js"></script>
  <style>
    *,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
    :root{
      --bg:#F3F4F6;--card:#FFFFFF;--border:#E5E7EB;
      --text:#0F172A;--text-sec:#64748B;--text-muted:#94A3B8;
      --primary:#2563EB;--primary-hover:#1D4ED8;--primary-light:#DBEAFE;
      --success:#10B981;--warning:#F59E0B;--danger:#EF4444;--danger-light:#FEE2E2;
      --radius:14px;--shadow:0 1px 3px rgba(0,0,0,.06),0 1px 2px rgba(0,0,0,.04);
      --shadow-lg:0 4px 6px -1px rgba(0,0,0,.07),0 2px 4px -2px rgba(0,0,0,.05);
    }
    body{font-family:'Inter',system-ui,-apple-system,sans-serif;background:var(--bg);color:var(--text);line-height:1.5;-webkit-font-smoothing:antialiased}
    .app{max-width:1400px;margin:0 auto;padding:20px 24px 40px}

    /* ─── Header ─── */
    .header{display:flex;align-items:center;justify-content:space-between;margin-bottom:24px;flex-wrap:wrap;gap:12px}
    .header h1{font-size:24px;font-weight:800;letter-spacing:-.5px;background:linear-gradient(135deg,var(--primary),#7C3AED);-webkit-background-clip:text;-webkit-text-fill-color:transparent}
    .header-right{display:flex;gap:8px;align-items:center}

    /* ─── Cards ─── */
    .card{background:var(--card);border-radius:var(--radius);border:1px solid var(--border);box-shadow:var(--shadow);padding:20px;transition:box-shadow .2s}
    .card:hover{box-shadow:var(--shadow-lg)}

    /* ─── KPI Row ─── */
    .kpi-row{display:grid;grid-template-columns:repeat(4,1fr);gap:16px;margin-bottom:20px}
    .kpi{position:relative;overflow:hidden}
    .kpi .kpi-icon{position:absolute;top:12px;right:12px;width:44px;height:44px;border-radius:12px;display:flex;align-items:center;justify-content:center;opacity:.12}
    .kpi .kpi-icon svg{width:28px;height:28px}
    .kpi-label{font-size:12px;font-weight:600;text-transform:uppercase;letter-spacing:.5px;color:var(--text-muted);margin-bottom:4px}
    .kpi-value{font-size:32px;font-weight:800;line-height:1.1;color:var(--text)}
    .kpi-sub{font-size:12px;color:var(--text-sec);margin-top:4px}

    /* ─── Main Grid ─── */
    .main-grid{display:grid;grid-template-columns:340px 1fr;gap:20px;align-items:start}

    /* ─── Form ─── */
    .form-card{position:sticky;top:20px}
    .form-title{font-size:16px;font-weight:700;margin-bottom:16px;display:flex;align-items:center;gap:8px}
    .form-title svg{width:20px;height:20px;color:var(--primary)}
    .field{position:relative;margin-bottom:14px}
    .field input,.field textarea,.field select{
      width:100%;padding:12px 14px;border:1.5px solid var(--border);border-radius:10px;
      font:inherit;font-size:14px;background:#FAFBFC;color:var(--text);
      transition:border-color .2s,box-shadow .2s;outline:none;resize:vertical
    }
    .field input:focus,.field textarea:focus,.field select:focus{border-color:var(--primary);box-shadow:0 0 0 3px var(--primary-light);background:#FFF}
    .field label{display:block;font-size:12px;font-weight:600;color:var(--text-sec);margin-bottom:5px;text-transform:uppercase;letter-spacing:.3px}
    .btn{display:inline-flex;align-items:center;justify-content:center;gap:6px;padding:10px 18px;border-radius:10px;font:inherit;font-size:14px;font-weight:600;cursor:pointer;border:none;transition:all .15s ease}
    .btn-primary{background:var(--primary);color:#FFF;width:100%}
    .btn-primary:hover{background:var(--primary-hover);transform:translateY(-1px);box-shadow:0 4px 12px rgba(37,99,235,.3)}
    .btn-secondary{background:#FFF;color:var(--text);border:1.5px solid var(--border)}
    .btn-secondary:hover{background:var(--bg);border-color:var(--primary);color:var(--primary)}
    .btn-ghost{background:transparent;color:var(--text-sec);padding:6px 10px;font-size:13px}
    .btn-ghost:hover{background:var(--bg);color:var(--text)}
    .btn-cancel{background:var(--bg);color:var(--text-sec);width:100%;margin-top:8px}
    .btn-cancel:hover{background:#E2E8F0}
    .status-msg{font-size:13px;margin-top:10px;padding:8px 12px;border-radius:8px;display:none}
    .status-msg.show{display:block}
    .status-msg.ok{background:#ECFDF5;color:#065F46}
    .status-msg.err{background:var(--danger-light);color:#991B1B}

    /* ─── Import Section ─── */
    .import-section{margin-top:16px;padding-top:16px;border-top:1px solid var(--border)}
    .import-section .field-label{font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.3px;color:var(--text-muted);margin-bottom:6px}
    .file-input-wrap{position:relative}
    .file-input-wrap input[type="file"]{width:100%;padding:10px;border:1.5px dashed var(--border);border-radius:10px;font:inherit;font-size:13px;background:#FAFBFC;cursor:pointer}
    .file-input-wrap input[type="file"]:hover{border-color:var(--primary)}
    .checkbox-label{display:flex;align-items:center;gap:8px;font-size:13px;color:var(--text-sec);margin:10px 0;cursor:pointer}
    .checkbox-label input[type="checkbox"]{width:16px;height:16px;accent-color:var(--primary)}
    .btn-import{background:var(--success);color:#FFF;width:100%}
    .btn-import:hover{background:#059669;transform:translateY(-1px)}

    /* ─── Filters ─── */
    .filter-bar{display:flex;align-items:end;gap:12px;flex-wrap:wrap;margin-bottom:16px}
    .filter-bar .field{margin-bottom:0;flex:1;min-width:140px}
    .filter-bar .field select{padding:10px 12px;font-size:13px}
    .filter-actions{display:flex;gap:8px}

    /* ─── Table ─── */
    .table-wrap{overflow-x:auto}
    table{width:100%;border-collapse:separate;border-spacing:0;font-size:13px}
    thead th{background:var(--bg);font-weight:600;font-size:11px;text-transform:uppercase;letter-spacing:.5px;color:var(--text-muted);padding:10px 14px;border-bottom:2px solid var(--border);text-align:left;white-space:nowrap}
    tbody td{padding:12px 14px;border-bottom:1px solid #F1F5F9;color:var(--text);vertical-align:middle}
    tbody tr{transition:background .15s}
    tbody tr:hover{background:#F8FAFC}
    .task-cell{max-width:200px;display:-webkit-box;-webkit-line-clamp:2;-webkit-box-orient:vertical;overflow:hidden;text-overflow:ellipsis;cursor:default;line-height:1.4}
    .task-cell:hover{-webkit-line-clamp:unset;overflow:visible}
    .action-btns{display:flex;gap:4px;white-space:nowrap}
    .action-btn{width:32px;height:32px;border-radius:8px;border:none;cursor:pointer;display:flex;align-items:center;justify-content:center;transition:all .15s}
    .action-btn svg{width:16px;height:16px}
    .action-btn.edit{background:var(--primary-light);color:var(--primary)}
    .action-btn.edit:hover{background:var(--primary);color:#FFF}
    .action-btn.del{background:var(--danger-light);color:var(--danger)}
    .action-btn.del:hover{background:var(--danger);color:#FFF}

    /* ─── Charts ─── */
    .charts-grid{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-top:16px}
    .chart-title{font-size:14px;font-weight:700;margin-bottom:12px;color:var(--text)}
    .chart-container{position:relative;height:260px}

    /* ─── History ─── */
    .history-section{margin-top:16px}
    .history-section h3{font-size:14px;font-weight:700;margin-bottom:12px}
    .history-section table{font-size:12px}
    .badge{display:inline-flex;padding:2px 8px;border-radius:6px;font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.3px}
    .badge-create{background:#ECFDF5;color:#065F46}
    .badge-update{background:#DBEAFE;color:#1E40AF}
    .badge-delete{background:var(--danger-light);color:#991B1B}
    .badge-import{background:#FEF3C7;color:#92400E}

    /* ─── Empty State ─── */
    .empty{text-align:center;padding:32px;color:var(--text-muted);font-size:14px}
    .empty svg{width:48px;height:48px;margin-bottom:8px;opacity:.3}

    /* ─── Responsive ─── */
    @media(max-width:1024px){
      .main-grid{grid-template-columns:1fr}
      .form-card{position:static}
      .kpi-row{grid-template-columns:repeat(2,1fr)}
      .charts-grid{grid-template-columns:1fr}
    }
    @media(max-width:640px){
      .kpi-row{grid-template-columns:1fr}
      .filter-bar{flex-direction:column}
      .header h1{font-size:20px}
    }
  </style>
</head>
<body>
<div class="app">
  <!-- Header -->
  <div class="header">
    <h1>⏱ ThuKudo Time Tracker</h1>
    <div class="header-right">
      <button class="btn btn-secondary" id="exportButton" title="Export Excel">
        <svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M21 15v4a2 2 0 01-2 2H5a2 2 0 01-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg>
        Export
      </button>
    </div>
  </div>

  <!-- KPI Cards -->
  <div class="kpi-row">
    <div class="card kpi" id="kpiHours">
      <div class="kpi-icon" style="background:var(--primary)"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><circle cx="12" cy="12" r="10"/><polyline points="12 6 12 12 16 14"/></svg></div>
      <div class="kpi-label">Monthly Hours</div>
      <div class="kpi-value" id="kpiHoursVal">0</div>
      <div class="kpi-sub">Total tracked time</div>
    </div>
    <div class="card kpi" id="kpiTasks">
      <div class="kpi-icon" style="background:var(--success)"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M9 11l3 3L22 4"/><path d="M21 12v7a2 2 0 01-2 2H5a2 2 0 01-2-2V5a2 2 0 012-2h11"/></svg></div>
      <div class="kpi-label">Entries</div>
      <div class="kpi-value" id="kpiTasksVal">0</div>
      <div class="kpi-sub">Work entries logged</div>
    </div>
    <div class="card kpi" id="kpiTopProject">
      <div class="kpi-icon" style="background:var(--warning)"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><polygon points="12 2 15.09 8.26 22 9.27 17 14.14 18.18 21.02 12 17.77 5.82 21.02 7 14.14 2 9.27 8.91 8.26 12 2"/></svg></div>
      <div class="kpi-label">Top Project</div>
      <div class="kpi-value" id="kpiTopVal" style="font-size:18px;word-break:break-word">—</div>
      <div class="kpi-sub">Most time invested</div>
    </div>
    <div class="card kpi" id="kpiAvg">
      <div class="kpi-icon" style="background:#7C3AED"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><line x1="18" y1="20" x2="18" y2="10"/><line x1="12" y1="20" x2="12" y2="4"/><line x1="6" y1="20" x2="6" y2="14"/></svg></div>
      <div class="kpi-label">Avg Hours/Day</div>
      <div class="kpi-value" id="kpiAvgVal">0</div>
      <div class="kpi-sub">Working days average</div>
    </div>
  </div>

  <!-- Main Grid -->
  <div class="main-grid">
    <!-- Left: Form -->
    <div class="card form-card">
      <div class="form-title" id="formTitle">
        <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><line x1="12" y1="5" x2="12" y2="19"/><line x1="5" y1="12" x2="19" y2="12"/></svg>
        Add Work Entry
      </div>
      <form id="entryForm">
        <div class="field"><label>Employee</label><input id="employee_name" name="employee_name" placeholder="Enter name..." required/></div>
        <div class="field"><label>Work Date</label><input id="work_date" name="work_date" type="date" required/></div>
        <div class="field"><label>Project</label><input id="project_name" name="project_name" placeholder="Project name..." required/></div>
        <div class="field"><label>Task</label><textarea id="task_name" name="task_name" rows="2" placeholder="What did you work on?" required></textarea></div>
        <div class="field"><label>Hours</label><input id="hours_spent" name="hours_spent" type="number" min="0.25" step="0.25" placeholder="0.00" required/></div>
        <div class="field"><label>Operator</label><input id="operator_name" name="operator_name" required/></div>
        <button class="btn btn-primary" id="submitButton" type="submit">
          <svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M19 21H5a2 2 0 01-2-2V5a2 2 0 012-2h11l5 5v11a2 2 0 01-2 2z"/><polyline points="17 21 17 13 7 13 7 21"/></svg>
          Save Entry
        </button>
        <button class="btn btn-cancel" id="cancelEditButton" type="button" style="display:none">Cancel Edit</button>
        <button class="btn btn-ghost" id="pinDefaultsBtn" type="button" style="margin-top:6px;font-size:12px;width:100%;border:1px dashed var(--border)" title="Save Employee/Project/Operator as defaults so they persist across sessions">
          <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M12 2L15 8.5L22 9.27L17 14.14L18.18 21.02L12 17.77L5.82 21.02L7 14.14L2 9.27L8.91 8.5L12 2z"/></svg>
          Save as Defaults
        </button>
        <div id="formStatus" class="status-msg"></div>
      </form>

      <!-- Import Section -->
      <div class="import-section">
        <div class="field-label">Import from Excel</div>
        <div class="file-input-wrap">
          <input id="importFile" type="file" accept=".xlsx"/>
        </div>
        <label class="checkbox-label">
          <input id="overwriteImport" type="checkbox" checked/>
          Overwrite duplicates
        </label>
        <button class="btn btn-import" id="importButton" type="button">
          <svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M21 15v4a2 2 0 01-2 2H5a2 2 0 01-2-2v-4"/><polyline points="17 8 12 3 7 8"/><line x1="12" y1="3" x2="12" y2="15"/></svg>
          Import Excel
        </button>
      </div>
    </div>

    <!-- Right: Data -->
    <div>
      <!-- Filters -->
      <div class="card" style="margin-bottom:16px">
        <div class="filter-bar">
          <div class="field"><label>Month</label><select id="monthFilter"></select></div>
          <div class="field"><label>Project</label><select id="projectFilter"></select></div>
          <div class="filter-actions">
            <button class="btn btn-secondary" id="refreshButton" type="button">
              <svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><polyline points="23 4 23 10 17 10"/><path d="M20.49 15a9 9 0 11-2.12-9.36L23 10"/></svg>
              Refresh
            </button>
          </div>
        </div>
      </div>

      <!-- Entries Table -->
      <div class="card" style="margin-bottom:16px">
        <h3 style="font-size:14px;font-weight:700;margin-bottom:12px">Work Entries</h3>
        <div class="table-wrap">
          <table>
            <thead><tr><th>Date</th><th>Employee</th><th>Project</th><th>Task</th><th>Hours</th><th style="width:80px">Actions</th></tr></thead>
            <tbody id="entriesBody"></tbody>
          </table>
        </div>
      </div>

      <!-- Charts -->
      <div class="charts-grid">
        <div class="card">
          <div class="chart-title">Hours by Project</div>
          <div class="chart-container"><canvas id="projectChart"></canvas></div>
        </div>
        <div class="card">
          <div class="chart-title">Hours by Employee</div>
          <div class="chart-container"><canvas id="employeeChart"></canvas></div>
        </div>
        <div class="card">
          <div class="chart-title">Daily Activity</div>
          <div class="chart-container"><canvas id="dailyChart"></canvas></div>
        </div>
        <div class="card">
          <div class="chart-title">Project Distribution</div>
          <div class="chart-container"><canvas id="projectPieChart"></canvas></div>
        </div>
      </div>

      <!-- History -->
      <div class="card history-section" style="margin-top:16px">
        <h3>Recent Changes</h3>
        <div class="table-wrap">
          <table>
            <thead><tr><th>Time</th><th>Action</th><th>By</th><th>Employee</th><th>Project</th><th>Date</th><th>Hours</th></tr></thead>
            <tbody id="historyBody"></tbody>
          </table>
        </div>
      </div>
    </div>
  </div>
</div>

<script>
const COLORS = ['#2563EB','#7C3AED','#10B981','#F59E0B','#EF4444','#06B6D4','#EC4899','#8B5CF6','#14B8A6','#F97316'];
const state = {month:'',project:'All Projects',editingId:null,entries:[]};
let projectChart,employeeChart,dailyChart,projectPieChart;

const $=id=>document.getElementById(id);
// Load saved defaults from localStorage
const savedDef=JSON.parse(localStorage.getItem('tk_defaults')||'{}');
$('employee_name').value=savedDef.employee_name||'';
$('project_name').value=savedDef.project_name||'';
$('operator_name').value=savedDef.operator_name||'__DEFAULT_OPERATOR__';
$('work_date').value=new Date().toISOString().slice(0,10);
function saveDefaults(){localStorage.setItem('tk_defaults',JSON.stringify({employee_name:$('employee_name').value,project_name:$('project_name').value,operator_name:$('operator_name').value}))}

function showStatus(msg,isError=false){
  const el=$('formStatus');
  el.textContent=msg;el.className='status-msg show '+(isError?'err':'ok');
  setTimeout(()=>{el.className='status-msg'},4000);
}
function softReset(){
  // Keep employee, project, operator from localStorage or current values
  const sd=JSON.parse(localStorage.getItem('tk_defaults')||'{}');
  state.editingId=null;
  $('task_name').value='';$('hours_spent').value='';
  $('work_date').value=new Date().toISOString().slice(0,10);
  $('employee_name').value=sd.employee_name||$('employee_name').value;
  $('project_name').value=sd.project_name||$('project_name').value;
  $('operator_name').value=sd.operator_name||$('operator_name').value;
  $('formTitle').innerHTML='<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" width="20" height="20"><line x1="12" y1="5" x2="12" y2="19"/><line x1="5" y1="12" x2="19" y2="12"/></svg> Add Work Entry';
  $('submitButton').innerHTML='<svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M19 21H5a2 2 0 01-2-2V5a2 2 0 012-2h11l5 5v11a2 2 0 01-2 2z"/><polyline points="17 21 17 13 7 13 7 21"/></svg> Save Entry';
  $('cancelEditButton').style.display='none';
  $('formStatus').className='status-msg';
}
function resetForm(){softReset()}

$('cancelEditButton').onclick=resetForm;
$('pinDefaultsBtn').onclick=()=>{saveDefaults();showStatus('Defaults saved! Employee/Project/Operator will be remembered.');};

$('entryForm').addEventListener('submit',async e=>{
  e.preventDefault();
  const fd=new FormData(e.target);const payload=Object.fromEntries(fd.entries());payload.hours_spent=Number(payload.hours_spent);
  if(!payload.operator_name?.trim()){showStatus('Please enter the operator name.',true);return}
  const isEdit=state.editingId!==null;
  if(isEdit&&!confirm('Update the selected entry?'))return;
  const resp=await fetch(isEdit?`/api/entries/${state.editingId}`:'/api/entries',{method:isEdit?'PUT':'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(payload)});
  const res=await resp.json();
  if(!resp.ok){showStatus(res.error||'Could not save.',true);return}
  showStatus(isEdit?'Entry updated successfully!':'Entry saved successfully!');
  saveDefaults(); // Save current employee/project/operator
  softReset();state.month=payload.work_date.slice(0,7);
  await initFilters();await loadDashboard();
});

$('refreshButton').onclick=loadDashboard;
$('exportButton').onclick=()=>{const q=new URLSearchParams({month:state.month,project:state.project});window.location.href=`/api/export.xlsx?${q}`};
$('importButton').onclick=async()=>{
  const op=$('operator_name').value.trim();const fileInput=$('importFile');const file=fileInput.files&&fileInput.files[0];
  if(!op){showStatus('Please enter operator name.',true);return}
  if(!file){showStatus('Please choose an Excel file.',true);return}
  const overwrite=$('overwriteImport').checked;
  if(!confirm(overwrite?'Import and overwrite duplicates?':'Import without overwriting?'))return;
  const fd=new FormData();fd.append('file',file);fd.append('operator_name',op);fd.append('overwrite',overwrite?'1':'0');
  const resp=await fetch('/api/import.xlsx',{method:'POST',body:fd});
  const res=await resp.json();
  if(!resp.ok){showStatus(res.error||'Import failed.',true);return}
  const r=res.result||{};
  showStatus(`Import done! Created: ${r.created||0}, Updated: ${r.updated||0}, Skipped: ${r.skipped||0}`);
  fileInput.value='';await initFilters();await loadDashboard();
};
$('monthFilter').onchange=e=>{state.month=e.target.value;loadDashboard()};
$('projectFilter').onchange=e=>{state.project=e.target.value;loadDashboard()};

async function initFilters(){
  const f=await(await fetch('/api/filters')).json();
  const months=f.months.length?f.months:[new Date().toISOString().slice(0,7)];
  const allMonths=['all',...months];
  const projects=f.projects.length?['All Projects',...f.projects]:['All Projects'];
  if(!state.month||(!allMonths.includes(state.month)&&state.month!=='all'))state.month=months[0]||'all';
  if(!projects.includes(state.project))state.project='All Projects';
  $('monthFilter').innerHTML=allMonths.map(x=>`<option value="${x}">${x==='all'?'\u2b50 All Months':x}</option>`).join('');
  $('projectFilter').innerHTML=projects.map(x=>`<option value="${x}">${x}</option>`).join('');
  $('monthFilter').value=state.month;$('projectFilter').value=state.project;
}

async function loadDashboard(){
  const q=new URLSearchParams({month:state.month,project:state.project});
  const data=await(await fetch(`/api/dashboard?${q}`)).json();
  state.entries=data.entries;
  renderKPI(data.summary);renderTable(data.entries);renderCharts(data.summary);renderHistory(data.history||[]);
}

function renderKPI(s){
  $('kpiHoursVal').textContent=s.total_hours.toFixed(1);
  $('kpiTasksVal').textContent=s.entry_count;
  const tp=s.top_project||'—';
  $('kpiTopVal').textContent=tp.length>30?tp.slice(0,30)+'…':tp;
  $('kpiTopVal').title=tp;
  const days=Object.keys(s.by_date||{}).length||1;
  $('kpiAvgVal').textContent=(s.total_hours/days).toFixed(1);
}

function renderTable(entries){
  const body=$('entriesBody');
  if(!entries.length){body.innerHTML='<tr><td colspan="6" class="empty"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.5"><path d="M9 17H5a2 2 0 01-2-2V5a2 2 0 012-2h14a2 2 0 012 2v3m-7 11l4-4m0 0l4 4m-4-4v12"/></svg><div>No entries for this period</div></td></tr>';return}
  body.innerHTML=entries.map(e=>`<tr>
    <td style="white-space:nowrap;font-weight:500">${e.work_date}</td>
    <td>${e.employee_name}</td>
    <td><span style="font-weight:500;color:var(--primary)">${e.project_name}</span></td>
    <td><div class="task-cell" title="${(e.task_name||'').replace(/"/g,'&quot;')}">${e.task_name}</div></td>
    <td style="font-weight:600">${e.hours_spent.toFixed(2)}</td>
    <td><div class="action-btns">
      <button class="action-btn edit" data-edit="${e.id}" title="Edit"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M17 3a2.83 2.83 0 114 4L7.5 20.5 2 22l1.5-5.5L17 3z"/></svg></button>
      <button class="action-btn del" data-del="${e.id}" title="Delete"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><polyline points="3 6 5 6 21 6"/><path d="M19 6v14a2 2 0 01-2 2H7a2 2 0 01-2-2V6m3 0V4a2 2 0 012-2h4a2 2 0 012 2v2"/></svg></button>
    </div></td>
  </tr>`).join('');
  body.querySelectorAll('[data-edit]').forEach(b=>b.onclick=()=>startEdit(+b.dataset.edit));
  body.querySelectorAll('[data-del]').forEach(b=>b.onclick=()=>deleteEntry(+b.dataset.del));
}

function startEdit(id){
  const e=state.entries.find(x=>x.id===id);if(!e)return;
  state.editingId=e.id;
  $('employee_name').value=e.employee_name;$('work_date').value=e.work_date;
  $('project_name').value=e.project_name;$('task_name').value=e.task_name;
  $('hours_spent').value=e.hours_spent;
  $('formTitle').innerHTML='<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" width="20" height="20"><path d="M17 3a2.83 2.83 0 114 4L7.5 20.5 2 22l1.5-5.5L17 3z"/></svg> Edit Entry #'+e.id;
  $('submitButton').innerHTML='<svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><polyline points="20 6 9 17 4 12"/></svg> Update Entry';
  $('cancelEditButton').style.display='block';
  showStatus('Editing entry #'+e.id);
  window.scrollTo({top:0,behavior:'smooth'});
}

async function deleteEntry(id){
  const op=$('operator_name').value.trim();
  if(!op){showStatus('Please enter operator name.',true);return}
  if(!confirm('Delete this entry?'))return;
  const resp=await fetch(`/api/entries/${id}`,{method:'DELETE',headers:{'Content-Type':'application/json'},body:JSON.stringify({operator_name:op})});
  const res=await resp.json();
  if(!resp.ok){showStatus(res.error||'Delete failed.',true);return}
  if(state.editingId===id)resetForm();
  showStatus('Entry deleted.');await loadDashboard();
}

function renderCharts(s){
  const projNames=Object.keys(s.by_project||{});
  const projVals=Object.values(s.by_project||{});
  const empNames=Object.keys(s.by_employee||{});
  const empVals=Object.values(s.by_employee||{});
  const dateKeys=Object.keys(s.by_date||{}).map(d=>d.slice(5));
  const dateVals=Object.values(s.by_date||{});

  // Bar Chart - Hours by Project
  if(projectChart)projectChart.destroy();
  projectChart=new Chart($('projectChart'),{type:'bar',data:{labels:projNames,datasets:[{label:'Hours',data:projVals,backgroundColor:COLORS.slice(0,projNames.length),borderRadius:6,borderSkipped:false}]},options:{responsive:true,maintainAspectRatio:false,indexAxis:'y',plugins:{legend:{display:false}},scales:{x:{grid:{display:false},ticks:{font:{size:11}}},y:{grid:{display:false},ticks:{font:{size:11,weight:500}}}}}});

  // Doughnut - Employee Share
  if(employeeChart)employeeChart.destroy();
  employeeChart=new Chart($('employeeChart'),{type:'doughnut',data:{labels:empNames,datasets:[{data:empVals,backgroundColor:COLORS.slice(0,empNames.length),borderWidth:2,borderColor:'#FFF',hoverOffset:8}]},options:{responsive:true,maintainAspectRatio:false,cutout:'65%',plugins:{legend:{position:'bottom',labels:{padding:12,usePointStyle:true,pointStyle:'circle',font:{size:11}}}}}});

  // Line Chart - Daily Activity
  if(dailyChart)dailyChart.destroy();
  dailyChart=new Chart($('dailyChart'),{type:'line',data:{labels:dateKeys,datasets:[{label:'Hours',data:dateVals,borderColor:'#2563EB',backgroundColor:'rgba(37,99,235,.08)',fill:true,tension:.4,pointBackgroundColor:'#2563EB',pointRadius:4,pointHoverRadius:6}]},options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},scales:{x:{grid:{display:false},ticks:{font:{size:10},maxRotation:45}},y:{beginAtZero:true,grid:{color:'#F1F5F9'},ticks:{font:{size:11}}}}}});

  // Pie - Project Distribution
  if(projectPieChart)projectPieChart.destroy();
  projectPieChart=new Chart($('projectPieChart'),{type:'pie',data:{labels:projNames,datasets:[{data:projVals,backgroundColor:COLORS.slice(0,projNames.length),borderWidth:2,borderColor:'#FFF'}]},options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'bottom',labels:{padding:12,usePointStyle:true,pointStyle:'circle',font:{size:11}}}}}});
}

function renderHistory(rows){
  const body=$('historyBody');
  if(!rows.length){body.innerHTML='<tr><td colspan="7" class="empty">No recent changes</td></tr>';return}
  const badgeClass=a=>{if(a.includes('DELETE'))return'badge-delete';if(a.includes('UPDATE')||a.includes('OVERWRITE'))return'badge-update';if(a.includes('IMPORT'))return'badge-import';return'badge-create'};
  body.innerHTML=rows.map(r=>`<tr>
    <td style="white-space:nowrap;font-size:11px;color:var(--text-muted)">${(r.changed_at||'').replace('T',' ').slice(0,19)}</td>
    <td><span class="badge ${badgeClass(r.action)}">${r.action}</span></td>
    <td>${r.changed_by||''}</td>
    <td>${r.employee_name||''}</td>
    <td>${r.project_name||''}</td>
    <td>${r.work_date||''}</td>
    <td style="font-weight:500">${Number(r.hours_spent||0).toFixed(2)}</td>
  </tr>`).join('');
}

(async()=>{await initFilters();await loadDashboard()})();
</script>
</body>
</html>"""
